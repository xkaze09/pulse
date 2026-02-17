"""Document ingestion pipeline for the Pulse RAG agent.

Loads PDFs and Excel files from data/documents/, enriches metadata with URLs
from url_map.json, chunks the text, embeds with OpenAI, and stores in ChromaDB.

Usage:
    cd backend
    python -m src.ingest
"""

import src.compat  # noqa: F401 — must be first to patch pydantic v1 for Python 3.14+

import json
import os
from datetime import datetime, timezone
from pathlib import Path

import chromadb
from chromadb.utils.embedding_functions import DefaultEmbeddingFunction
from langchain_community.document_loaders import PyPDFLoader
from langchain_core.documents import Document
from langchain_text_splitters import RecursiveCharacterTextSplitter

from src.config import (
    CHROMA_PERSIST_DIR,
    COLLECTION_NAME,
    DOCUMENTS_DIR,
    ORG_DIR,
    URL_MAP_PATH,
)


def load_url_map(url_map_path: Path) -> dict[str, str]:
    """Load the filename -> public URL mapping from JSON."""
    if not url_map_path.exists():
        print(f"Warning: URL map not found at {url_map_path}. URLs will not be attached.")
        return {}
    with open(url_map_path, "r", encoding="utf-8") as f:
        return json.load(f)


def _load_excel(file_path: Path) -> list[Document]:
    """Load an Excel file (.xlsx / .xls) into LangChain Documents.

    Each sheet becomes one or more Documents. Row data is converted to readable
    text so the content is searchable and embeddable.
    """
    import openpyxl

    docs: list[Document] = []
    wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Use first row as headers if it looks like a header row
        headers = [str(c) if c is not None else f"Col{i+1}" for i, c in enumerate(rows[0])]
        text_lines: list[str] = [f"Sheet: {sheet_name}", ""]

        for row_idx, row in enumerate(rows[1:], start=2):
            parts: list[str] = []
            for h, val in zip(headers, row):
                if val is not None:
                    parts.append(f"{h}: {val}")
            if parts:
                text_lines.append(f"Row {row_idx}: " + " | ".join(parts))

        if len(text_lines) > 2:  # more than just the header
            docs.append(
                Document(
                    page_content="\n".join(text_lines),
                    metadata={
                        "source": str(file_path),
                        "sheet": sheet_name,
                        "page": 0,
                    },
                )
            )

    wb.close()
    return docs


def _load_markdown(file_path: Path) -> list[Document]:
    """Load a Markdown file (.md) into LangChain Documents.

    Splits on heading boundaries (## / ###) so each section becomes its own
    searchable chunk, preserving heading context in the text.
    """
    text = file_path.read_text(encoding="utf-8")
    import re

    # Split on H2/H3 headings, keeping the heading in each section
    sections = re.split(r"(?=^#{1,3} )", text, flags=re.MULTILINE)
    docs: list[Document] = []
    for i, section in enumerate(sections):
        content = section.strip()
        if not content:
            continue
        docs.append(
            Document(
                page_content=content,
                metadata={
                    "source": str(file_path),
                    "page": i,
                },
            )
        )
    return docs if docs else [Document(page_content=text, metadata={"source": str(file_path), "page": 0})]


def _load_org_json(file_path: Path) -> list[Document]:
    """Convert an org/process/workflow JSON file into LangChain Documents.

    Each JSON becomes two Documents:
      1. A summary of all nodes (label + description) — great for keyword search.
      2. A step-by-step sequence following the edge flow — great for process questions.
    """
    with open(file_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    diagram_type = data.get("diagram_type", file_path.stem)
    nodes: list[dict] = data.get("nodes", [])
    edges: list[dict] = data.get("edges", [])

    # Build id → node lookup
    node_map = {n["id"]: n for n in nodes}

    # --- Document 1: Node descriptions ---
    type_label = diagram_type.replace("_", " ").title()
    lines = [f"# {type_label}", ""]
    for node in nodes:
        label = node.get("label", "")
        desc = node.get("description", "")
        node_type = node.get("node_type", "")
        permission = node.get("permission_level", "public")
        lines.append(f"## {label}")
        if desc:
            lines.append(desc)
        lines.append(f"Type: {node_type} | Access: {permission}")
        lines.append("")

    doc_nodes = Document(
        page_content="\n".join(lines),
        metadata={
            "source": str(file_path),
            "filename": file_path.name,
            "diagram_type": diagram_type,
            "section": "nodes",
            "page": 0,
            "url": "",
        },
    )

    # --- Document 2: Step-by-step flow ---
    flow_lines = [f"# {type_label} — Step-by-Step Flow", ""]
    step = 1
    for edge in edges:
        src = node_map.get(edge.get("source_id", ""), {})
        tgt = node_map.get(edge.get("target_id", ""), {})
        src_label = src.get("label", edge.get("source_id", ""))
        tgt_label = tgt.get("label", edge.get("target_id", ""))
        edge_label = edge.get("label", "")
        connector = f' ({edge_label})' if edge_label else ""
        flow_lines.append(f"Step {step}: {src_label}{connector} → {tgt_label}")
        src_desc = src.get("description", "")
        if src_desc:
            flow_lines.append(f"  {src_desc}")
        step += 1
    # Add description of last node (terminal)
    if edges:
        last_tgt = node_map.get(edges[-1].get("target_id", ""), {})
        if last_tgt.get("description"):
            flow_lines.append(f"  {last_tgt['description']}")

    doc_flow = Document(
        page_content="\n".join(flow_lines),
        metadata={
            "source": str(file_path),
            "filename": file_path.name,
            "diagram_type": diagram_type,
            "section": "flow",
            "page": 1,
            "url": "",
        },
    )

    return [doc_nodes, doc_flow]


def load_org_diagrams(org_dir: Path) -> list[Document]:
    """Load all org/process/workflow JSON files from data/org/ as Documents."""
    if not org_dir.exists():
        print(f"  Org directory not found: {org_dir}")
        return []

    all_docs: list[Document] = []
    json_files = sorted(org_dir.glob("*.json"))

    if not json_files:
        print(f"  No JSON files found in {org_dir}")
        return all_docs

    for file_path in json_files:
        print(f"Loading: {file_path.name}")
        docs = _load_org_json(file_path)
        all_docs.extend(docs)
        print(f"  Loaded {len(docs)} section(s) from {file_path.name}")

    return all_docs


def load_documents(documents_dir: Path, url_map: dict[str, str]) -> list:
    """Load all PDFs, Excel, and Markdown files from the documents directory with enriched metadata."""
    all_docs = []

    # Collect supported files
    supported_extensions = (".pdf", ".xlsx", ".xls", ".md")
    files = [
        f for f in documents_dir.iterdir()
        if f.is_file() and f.suffix.lower() in supported_extensions
    ]

    if not files:
        print(f"No supported files found in {documents_dir}")
        print(f"  Supported formats: {', '.join(supported_extensions)}")
        return all_docs

    for file_path in sorted(files):
        print(f"Loading: {file_path.name}")
        ext = file_path.suffix.lower()

        if ext == ".pdf":
            loader = PyPDFLoader(str(file_path))
            docs = loader.load()
        elif ext in (".xlsx", ".xls"):
            docs = _load_excel(file_path)
        elif ext == ".md":
            docs = _load_markdown(file_path)
        else:
            continue

        # Enrich metadata
        filename = file_path.name
        public_url = url_map.get(filename, "")

        for doc in docs:
            doc.metadata["url"] = public_url
            doc.metadata["filename"] = filename
            doc.metadata["last_updated"] = datetime.now(timezone.utc).strftime("%Y-%m-%d")
            if "page" not in doc.metadata:
                doc.metadata["page"] = 0

        all_docs.extend(docs)
        print(f"  Loaded {len(docs)} section(s) from {filename}")

    return all_docs


def chunk_documents(documents: list) -> list:
    """Split documents into chunks using character-based splitting."""
    splitter = RecursiveCharacterTextSplitter(
        chunk_size=4000,
        chunk_overlap=400,
    )
    chunks = splitter.split_documents(documents)
    print(f"Split {len(documents)} pages into {len(chunks)} chunks")
    return chunks


def create_vector_store(chunks: list) -> None:
    """Create and persist a ChromaDB vector store from document chunks."""
    persist_dir = CHROMA_PERSIST_DIR

    # Remove old store if it exists
    if os.path.exists(persist_dir):
        import shutil
        shutil.rmtree(persist_dir)
        print(f"Removed existing vector store at {persist_dir}")

    client = chromadb.PersistentClient(path=persist_dir)
    collection = client.get_or_create_collection(
        name=COLLECTION_NAME,
        embedding_function=DefaultEmbeddingFunction(),
    )

    ids = [str(i) for i in range(len(chunks))]
    documents = [c.page_content for c in chunks]
    metadatas = [c.metadata for c in chunks]

    # Add in batches of 100 to avoid memory issues
    batch = 100
    for start in range(0, len(chunks), batch):
        collection.add(
            ids=ids[start : start + batch],
            documents=documents[start : start + batch],
            metadatas=metadatas[start : start + batch],
        )

    print(f"Created vector store with {len(chunks)} chunks in '{COLLECTION_NAME}'")
    print(f"Persisted to: {persist_dir}")


def main():
    """Run the full ingestion pipeline."""
    print("=" * 60)
    print("Pulse - Document Ingestion Pipeline")
    print("=" * 60)

    # Step 1: Load URL map
    print("\n[1/4] Loading URL map...")
    url_map = load_url_map(URL_MAP_PATH)
    print(f"  Found {len(url_map)} URL mappings")

    # Step 2: Load documents
    print("\n[2/4] Loading documents...")
    documents = load_documents(DOCUMENTS_DIR, url_map)

    print("\n      Loading org/process/workflow diagrams...")
    org_docs = load_org_diagrams(ORG_DIR)
    documents.extend(org_docs)

    if not documents:
        print("No documents to ingest. Exiting.")
        return

    # Step 3: Chunk documents
    print("\n[3/4] Chunking documents...")
    chunks = chunk_documents(documents)

    # Step 4: Create vector store
    print("\n[4/4] Creating vector store...")
    create_vector_store(chunks)

    print("\n" + "=" * 60)
    print("Ingestion complete!")
    print("=" * 60)


if __name__ == "__main__":
    main()
